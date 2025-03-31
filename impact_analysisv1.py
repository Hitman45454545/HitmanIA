import streamlit as st
import pandas as pd
from openpyxl import Workbook

def validation (uploaded_file, tab_name, sort_order, month_col, rate_col_1, rate_col_2, selected_month,
                monitored_flag_col, use_monitored, region_col, selected_region, r_n_ind, adequacy_col):

    # Read the Excel file
    df = pd.read_excel(uploaded_file, sheet_name=tab_name, header=1)

    # Filter data by selected region
    df = df[df[region_col] == selected_region]
    df_all = df[(df[monitored_flag_col].isin([0, 1])) & (df[r_n_ind] == "R")]
    df_unmon = df_all[(df_all[monitored_flag_col] == 0)]
    df_all[rate_col_2] = pd.to_numeric(df_all[rate_col_2], errors='coerce')
    df_all[month_col] = pd.to_numeric(df_all[month_col], errors="coerce")
    df_filtered_all = df_all[df_all[month_col] == selected_month]

    # Filter by monitored flag if applicable
    if use_monitored and monitored_flag_col:
        if monitored_flag_col in df.columns:
            df = df[(df[monitored_flag_col] == 1) & (df[r_n_ind] =="R")]  # Filter only monitored data and renewal data for monitored calculation.
        else:
            st.warning(f"Monitored flag column '{monitored_flag_col}' not found. Proceeding without filtering.")


    # Ensure required columns exist
    required_columns = [month_col, rate_col_1, rate_col_2, region_col, adequacy_col]
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        st.error(f"Missing required columns: {', '.join(missing_cols)}")
    invalid_rows_data = {}  # Store rows failing validation checks

    # ✅ Validate Adequacy Column
    invalid_adequacy = df[~df[adequacy_col].apply(lambda x: pd.isna(x) or isinstance(x, (int, float)))]
    if not invalid_adequacy.empty:
        invalid_rows_data["Adequacy Column Issue"] = invalid_adequacy

    # ✅ Validate Rate_col_1 and New/Renewal Flag
    invalid_rate_rn = df[(df[rate_col_1].notna()) & ((df[r_n_ind] != "R") | (df[monitored_flag_col] != 1))]
    if not invalid_rate_rn.empty:
        invalid_rows_data["Rate Change & New/Renewal Flag Issue"] = invalid_rate_rn

    # ✅ Validate Inception Month
    invalid_month = df[~df[month_col].apply(lambda x: pd.isna(x) or (isinstance(x, (int, float)) and x <= selected_month))]
    if not invalid_month.empty:
        invalid_rows_data["Inception Month Issue"] = invalid_month

    # Create a new Excel file
    output_file = "impact_analysis.xlsx"
    wb = Workbook()

    # Tab 1: Validation
    ws1 = wb.create_sheet(title="Validation")
    # Iterate over each issue and its corresponding DataFrame
    for issue, df_problem in invalid_rows_data.items():
        ws1.append([issue])  # Add issue type as a separate row
        if not df_problem.empty:
            ws1.append(df_problem.columns.tolist())  # Column headers
            for row in df_problem.itertuples(index=False):
                ws1.append(list(row))  # Add data rows
        ws1.append([])  # Empty row for spacing between issues
  

    # Save the workbook
    wb.save(output_file)

    return (output_file, invalid_rows_data)

def process_excel(uploaded_file, tab_name, sort_order, month_col, rate_col_1, rate_col_2, selected_month,
                  monitored_flag_col, use_monitored, region_col, selected_region, r_n_ind, adequacy_col):
    # Read the Excel file
    df = pd.read_excel(uploaded_file, sheet_name=tab_name, header=1)

    # Filter data by selected region
    df = df[df[region_col] == selected_region]
    df_all = df[(df[monitored_flag_col].isin([0, 1])) & (df[r_n_ind] == "R")]
    df_unmon = df_all[(df_all[monitored_flag_col] == 0)]
    df_all[rate_col_2] = pd.to_numeric(df_all[rate_col_2], errors='coerce')
    df_all[month_col] = pd.to_numeric(df_all[month_col], errors="coerce")
    df_filtered_all = df_all[df_all[month_col] == selected_month]

    # Filter by monitored flag if applicable
    if use_monitored and monitored_flag_col:
        if monitored_flag_col in df.columns:
            df = df[(df[monitored_flag_col] == 1) & (df[r_n_ind] =="R")]  # Filter only monitored data and renewal data for monitored calculation.
        else:
            st.warning(f"Monitored flag column '{monitored_flag_col}' not found. Proceeding without filtering.")


    # Ensure required columns exist
    required_columns = [month_col, rate_col_1, rate_col_2, region_col, adequacy_col]
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        st.error(f"Missing required columns: {', '.join(missing_cols)}")
    invalid_rows_data = {}  # Store rows failing validation checks

    # Compute "Hypothetical Premium"
    df[rate_col_1] = pd.to_numeric(df[rate_col_1], errors='coerce')
    df[rate_col_2] = pd.to_numeric(df[rate_col_2], errors='coerce')
    df["HypotheticalPremium"] = df[rate_col_2] / (1 + df[rate_col_1])

    # Compute "Rate Change Amount"
    df["RateChangeAmount"] = df[rate_col_1] * df["HypotheticalPremium"]
    df[month_col] = pd.to_numeric(df[month_col], errors="coerce")

    # Filter data by selected month
    df_filtered = df[df[month_col] == selected_month]

    # Sort data
    ascending = sort_order == "Lowest to Highest"
    df_sorted = df_filtered.sort_values(by="RateChangeAmount", ascending=ascending)

    # Compute YTD and Monthly Rate Change
    ytd_rate_change = df["RateChangeAmount"].sum() / df["HypotheticalPremium"].sum() if not df.empty else 0
    monthly_rate_change = df_filtered["RateChangeAmount"].sum() / df_filtered["HypotheticalPremium"].sum() if not df_filtered.empty else 0

    # Remove Top 10 from Selected Month
    df_excluded_month = df_sorted.iloc[driver_count:]
    top_10_policies = df_sorted.iloc[:driver_count]
    top_10_values = top_10_policies[["RateChangeAmount", "HypotheticalPremium"]]

    top_10_policies_values = df_sorted.iloc[:driver_count]["RateChangeAmount"].values
    df_excluded_ytd = df[~df["RateChangeAmount"].isin(top_10_policies_values)]

    # Compute Rate Change after exclusions
    ytd_rate_change_excluded = df_excluded_ytd["RateChangeAmount"].sum() / df_excluded_ytd["HypotheticalPremium"].sum() if not df_excluded_ytd.empty else 0
    monthly_rate_change_excluded = df_excluded_month["RateChangeAmount"].sum() / df_excluded_month["HypotheticalPremium"].sum() if not df_excluded_month.empty else 0

    # Sort data for un mont policies
    df_unmon_sorted = df_unmon.sort_values(by=rate_col_2, ascending=False)
    df_unmon_excluded = df_unmon_sorted.iloc[unmon_driver_count:]
    top_10_policies_unmon = df_unmon_sorted.iloc[:unmon_driver_count]
    top_10_values_unmon = top_10_policies_unmon[rate_col_2]

    top_10_policies_values_unmon = df_unmon_sorted.iloc[:unmon_driver_count][rate_col_2].values
    df_excluded_ytd_unmon = df_all[~df_all[rate_col_2].isin(top_10_policies_values_unmon)]

    # Compute Rate Change after exclusions
    Ytd_monitored_percentage_after_exclusion = df[rate_col_2].sum() / df_excluded_ytd_unmon[rate_col_2].sum()

    # Compute monitored percentage
    Ytd_monitored_percentage = df[rate_col_2].sum() / df_all[rate_col_2].sum()
    monthly_monitored_percentage = df_filtered[rate_col_2].sum() / df_filtered_all[rate_col_2].sum()

    # Create a new Excel file
    output_file = "impact_analysis.xlsx"
    wb = Workbook()

    # Tab 1: Overall Rate Change Summary
    ws1 = wb.active
    ws1.title = "Rate Change Summary"
    ws1.append(["Metric", "Value"])
    ws1.append(["YTD Rate Change", ytd_rate_change])
    ws1.append(["Monthly Rate Change", monthly_rate_change])

    # Tab 2: Data Excluding Top 10 of Selected Month
    ws2 = wb.create_sheet(title="Data Excluding Drivers")
    ws2.append(["Metric", "Value"])
    ws2.append(["YTD Rate Change (Excl. Drivers of Month)", ytd_rate_change_excluded])
    ws2.append(["Monthly Rate Change (Excl. Drivers of Month)", monthly_rate_change_excluded])
    ws2.append([])  # Empty row
    ws2.append(["YTD Data Excluding Drivers of Selected Month"])
    ws2.append(df_excluded_ytd.columns.tolist())
    for row in df_excluded_ytd.itertuples(index=False):
        ws2.append(list(row))

    ws2.append([])  # Empty row
    ws2.append(["Monthly Data Excluding Drivers of Selected Month"])
    ws2.append(df_excluded_month.columns.tolist())
    for row in df_excluded_month.itertuples(index=False):
        ws2.append(list(row))

    # Tab 3: Top Policies of Selected Month
    ws3 = wb.create_sheet(title="Driver Policies")
    df_top10 = df_sorted.iloc[:driver_count]
    ws3.append(df_top10.columns.tolist())
    for row in df_top10.itertuples(index=False):
        ws3.append(list(row))

    # Tab 4: Top Policies unmonitored
    ws4 = wb.create_sheet(title="Un-monitored drivers")
    df_top10_unmon = df_unmon_sorted.iloc[:unmon_driver_count]
    ws4.append(df_top10_unmon.columns.tolist())
    for row in df_top10_unmon.itertuples(index=False):
        ws4.append(list(row))


    # Save the workbook
    wb.save(output_file)

    return (output_file, top_10_policies_unmon, top_10_policies, ytd_rate_change, monthly_rate_change,
            ytd_rate_change_excluded, monthly_rate_change_excluded, monthly_monitored_percentage,
            Ytd_monitored_percentage, Ytd_monitored_percentage_after_exclusion, invalid_rows_data)


# Set page config
st.set_page_config(page_title="Impact Analysis Tool", layout="wide")
st.title("Impact Analysis Tool")

# Sidebar for navigation
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["Rate Change Metrics", "Validation"])

# File uploader
with st.sidebar:
    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"], label_visibility="collapsed")
    st.markdown("### Settings")
    tab_name = st.selectbox("Enter Sheet Name", ["EMEA Fin Lines", "EMEA Casualty", "EMEA and LATAM B&M", "EMEA Marine", "EMEA Terrorism", "EMEA Tech", "EMEA Fire", "CEMENA Energy", "EMEA Environmental", "UKISA A&H", "Custom"])
    st.info("Below parameters can be customized. You can change them according to your needs.")
    month_col = st.text_input("Month Column Name ", value="Inception Month")
    rate_col_1 = st.text_input("Rate Change Amount Column ", value="Rate Change")
    rate_col_2 = st.text_input("Total Premium or Renewal Premium Column ", value="Total Premium")
    region_col = st.text_input("Region Column Name ", value="Region")
    monitored_flag_col = st.text_input("Monitored Flag Column ", value="Monitored Flag")
    selected_region = st.selectbox("Select Region", ["UKISA", "CEMENA"])
    use_monitored = st.radio("Use Monitored Flag for Calculations?\n(For validation function use no and for Rate change and monitored percentage change metrics select yess)", ["No", "Yes"])
    sort_order = st.selectbox("Sort Order", ["Highest to Lowest", "Lowest to Highest"])
    selected_month = st.selectbox("Select Month (Numeric)", list(range(1, 13)))
    driver_count = st.number_input("Enter the number of drivers:", value=10, min_value=0)
    r_n_ind = st.text_input("Renewal or New flag Column Name ", value="New/Renewal Flag")
    unmon_driver_count = st.number_input("Enter the number of unmonitored drivers:", value=10, min_value=0)
    adequacy_col = st.text_input("Adequacy Column Name ", value="Adequacy")

if tab_name == "Custom":
    custom_input = st.text_input("Please enter your custom Sheet Name:")
    if custom_input:
        tab_name = custom_input  # Update the tab_name with the user's custom input
else:
    custom_input = None  # Reset custom_input if a predefined option is selected
    
    # Process data only when clicking a specific button
if page == "Rate Change Metrics":
    if st.button("Get Metrics Summary"):
        col1, col2, col3 = st.columns(3)  # Create three columns for structured layout

        output_path, top_10_policies_unmon, top_10_policies, ytd_rate_change, monthly_rate_change, ytd_rate_change_excluded, monthly_rate_change_excluded, monthly_monitored_percentage, Ytd_monitored_percentage, Ytd_monitored_percentage_after_exclusion, invalid_rows_data = process_excel(
            uploaded_file, tab_name, sort_order, month_col, rate_col_1, rate_col_2,
            selected_month, monitored_flag_col, use_monitored == "Yes", region_col, selected_region, r_n_ind, adequacy_col
        )

        with col1:
            st.subheader(f"Top {driver_count} Policies")
            if top_10_policies is not None:
                st.dataframe(top_10_policies, use_container_width=True)

        with col2:
            st.subheader(f"Top {unmon_driver_count} Policies Unmonitored")
            if top_10_policies_unmon is not None:
                st.dataframe(top_10_policies_unmon, use_container_width=True)

        with col3:
            st.subheader("Rate Change Metrics")
            # Convert the decimal rates to percentage format
            st.markdown(f"**YTD Rate Change:** {ytd_rate_change * 100:.2f}%")
            st.markdown(f"**Monthly Rate Change:** {monthly_rate_change * 100:.2f}%")
            st.markdown(f"**YTD Rate Change (Excl. Drivers):** {ytd_rate_change_excluded * 100:.2f}%")
            st.markdown(f"**Monthly Rate Change (Excl. Drivers):** {monthly_rate_change_excluded * 100:.2f}%")
            st.markdown(f"**Monthly Monitored Percentage:** {monthly_monitored_percentage * 100:.2f}%")
            st.markdown(f"**YTD Monitored Percentage:** {Ytd_monitored_percentage * 100:.2f}%")
            st.markdown(f"**YTD Monitored Percentage after Exclusion:** {Ytd_monitored_percentage_after_exclusion * 100:.2f}%")

        if output_path:
            with open(output_path, "rb") as file:
                st.download_button(label="Download Processed Excel", data=file, file_name="impact_analysis.xlsx")

           

 
elif page == "Validation":
    if st.button("Do validation"):
        output_path, invalid_rows_data = validation(
            uploaded_file, tab_name, sort_order, month_col, rate_col_1, rate_col_2,
            selected_month, monitored_flag_col, use_monitored == "Yes", region_col, selected_region, r_n_ind, adequacy_col
        )

         # ✅ Display Invalid Rows at the End
        if invalid_rows_data:
            st.warning("⚠️ Validation issues detected! Check the problematic rows below:")
            for issue, df_problem in invalid_rows_data.items():
                st.subheader(issue)
                st.dataframe(df_problem)
        if output_path:
            with open(output_path, "rb") as file:
                st.download_button(label="Download Processed Excel", data=file, file_name="impact_analysis.xlsx")


